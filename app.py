"""
SIP Tracker Dashboard — Groww Mutual Fund Analyzer
Supports both Groww Excel import and manual SIP entry.
Google SSO for access control.
"""

import io
import json
import os
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from functools import wraps

from dotenv import load_dotenv
load_dotenv()

import numpy as np
import openpyxl
import requests
from authlib.integrations.flask_client import OAuth
from flask import Flask, jsonify, render_template, request, redirect, url_for, session, abort
from scipy.optimize import brentq

app = Flask(__name__)
app.config['TEMPLATES_AUTO_RELOAD'] = True
app.secret_key = os.getenv('FLASK_SECRET_KEY', 'dev-secret-change-me')

# ---------------------------------------------------------------------------
# Google OAuth
# ---------------------------------------------------------------------------
GOOGLE_CLIENT_ID     = os.getenv('GOOGLE_CLIENT_ID', '')
GOOGLE_CLIENT_SECRET = os.getenv('GOOGLE_CLIENT_SECRET', '')
ALLOWED_EMAILS       = {e.strip().lower() for e in os.getenv('ALLOWED_EMAIL', '').split(',') if e.strip()}
OAUTH_CONFIGURED     = bool(GOOGLE_CLIENT_ID and GOOGLE_CLIENT_SECRET)

oauth = OAuth(app)
if OAUTH_CONFIGURED:
    oauth.register(
        name='google',
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
        server_metadata_url='https://accounts.google.com/.well-known/openid-configuration',
        client_kwargs={'scope': 'openid email profile'},
    )


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not OAUTH_CONFIGURED:
            return f(*args, **kwargs)          # SSO not set up yet — allow through
        if 'user' not in session:
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


@app.route('/login')
def login_page():
    if not OAUTH_CONFIGURED:
        return render_template('login.html', setup_needed=True)
    if 'user' in session:
        return redirect('/')
    return render_template('login.html', setup_needed=False)


@app.route('/auth/google')
def auth_google():
    if not OAUTH_CONFIGURED:
        return redirect('/')
    redirect_uri = url_for('auth_callback', _external=True)
    return oauth.google.authorize_redirect(redirect_uri)


@app.route('/auth/callback')
def auth_callback():
    if not OAUTH_CONFIGURED:
        return redirect('/')
    try:
        token    = oauth.google.authorize_access_token()
        userinfo = token.get('userinfo') or oauth.google.userinfo()
    except Exception as e:
        return render_template('login.html', setup_needed=False,
                               error=f"Google login failed: {e}")

    email = (userinfo.get('email') or '').lower().strip()

    if ALLOWED_EMAILS and email not in ALLOWED_EMAILS:
        session.clear()
        return render_template('login.html', setup_needed=False,
                               error=f"Access denied. This dashboard is private.")

    session['user'] = {
        'email':   email,
        'name':    userinfo.get('name', email),
        'picture': userinfo.get('picture', ''),
    }
    return redirect('/')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
DATA_DIR        = os.path.join(BASE_DIR, 'data', 'users')   # local dev only
CACHE_DIR       = os.path.join(BASE_DIR, 'cache')           # local dev only
CACHE_TTL       = 6 * 3600   # 6 hours NAV cache
STOCK_CACHE_TTL = 15 * 60    # 15 minutes stock price cache
OWNER_EMAIL     = os.getenv('OWNER_EMAIL', '').lower().strip()

# ---------------------------------------------------------------------------
# Supabase (production) vs local JSON (development)
# ---------------------------------------------------------------------------
SUPABASE_URL = os.getenv('SUPABASE_URL', '')
# Use the service_role key (server-side only — bypasses RLS safely)
SUPABASE_KEY = os.getenv('SUPABASE_SERVICE_KEY', '')
USE_SUPABASE = bool(SUPABASE_URL and SUPABASE_KEY)

_sb = None
if USE_SUPABASE:
    from supabase import create_client
    _sb = create_client(SUPABASE_URL, SUPABASE_KEY)


def _get_email():
    return (session.get('user') or {}).get('email', 'default')

# ---------------------------------------------------------------------------
# Benchmark XIRR targets by category
# ---------------------------------------------------------------------------
CATEGORY_BENCHMARKS = {
    "large cap":          {"xirr": 12.0, "index": "Nifty 50"},
    "large & mid cap":    {"xirr": 13.5, "index": "Nifty LargeMidcap 250"},
    "mid cap":            {"xirr": 15.0, "index": "Nifty Midcap 150"},
    "small cap":          {"xirr": 16.0, "index": "Nifty Smallcap 250"},
    "flexi cap":          {"xirr": 13.0, "index": "Nifty 500"},
    "multi cap":          {"xirr": 13.0, "index": "Nifty 500"},
    "focused":            {"xirr": 13.0, "index": "Nifty 500"},
    "elss":               {"xirr": 13.0, "index": "Nifty 500"},
    "value":              {"xirr": 13.0, "index": "Nifty 500 Value 50"},
    "value oriented":     {"xirr": 13.0, "index": "Nifty 500 Value 50"},
    "contra":             {"xirr": 13.0, "index": "Nifty 500"},
    "dividend yield":     {"xirr": 12.0, "index": "Nifty Dividend Opportunities 50"},
    "sectoral/thematic":  {"xirr": 12.0, "index": "Nifty 500"},
    "thematic":           {"xirr": 12.0, "index": "Nifty 500"},
    "sectoral":           {"xirr": 12.0, "index": "Nifty 500"},
    "hybrid":             {"xirr": 10.5, "index": "Crisil Hybrid 50+50 Moderate"},
    "balanced advantage": {"xirr": 10.0, "index": "Crisil Hybrid 50+50 Moderate"},
    "dynamic asset allocation": {"xirr": 10.0, "index": "Crisil Hybrid 50+50 Moderate"},
    "aggressive hybrid":  {"xirr": 11.0, "index": "Nifty 50 Hybrid Composite Debt 65:35"},
    "arbitrage":          {"xirr": 6.5,  "index": "Crisil Liquid Fund"},
    "debt":               {"xirr": 7.0,  "index": "Crisil Composite Bond"},
    "short duration":     {"xirr": 7.0,  "index": "Crisil Short Term Bond"},
    "liquid":             {"xirr": 6.5,  "index": "Crisil Liquid Fund"},
    "overnight":          {"xirr": 6.0,  "index": "Crisil Overnight"},
    "index":              {"xirr": 12.0, "index": "Nifty 50"},
    "etf":                {"xirr": 12.0, "index": "Nifty 50"},
    "gold":               {"xirr": 8.0,  "index": "Gold Price INR"},
    "international":      {"xirr": 10.0, "index": "MSCI World"},
    "equity":             {"xirr": 13.0, "index": "Nifty 500"},
}

# Groww sub-category → our category key
SUBCATEGORY_MAP = {
    "small cap":                  "small cap",
    "large cap":                  "large cap",
    "large & midcap":             "large & mid cap",
    "large & mid cap":            "large & mid cap",
    "mid cap":                    "mid cap",
    "elss":                       "elss",
    "thematic":                   "thematic",
    "sectoral":                   "sectoral/thematic",
    "international":              "international",
    "dynamic asset allocation":   "dynamic asset allocation",
    "aggressive hybrid":          "aggressive hybrid",
    "short duration":             "short duration",
    "value oriented":             "value oriented",
    "flexi cap":                  "flexi cap",
    "multi cap":                  "multi cap",
    "focused":                    "focused",
    "contra":                     "contra",
    "dividend yield":             "dividend yield",
    "overnight":                  "overnight",
    "liquid":                     "liquid",
    "ultra short duration":       "debt",
    "low duration":               "debt",
    "medium duration":            "debt",
    "conservative hybrid":        "hybrid",
    "arbitrage":                  "arbitrage",
    "index funds":                "index",
    "other etfs":                 "etf",
    "gold":                       "gold",
    "fund of funds investing overseas": "international",
}


# ---------------------------------------------------------------------------
# Data helpers — Supabase in production, JSON files locally
# ---------------------------------------------------------------------------

def _local_sips_path():
    safe = _get_email().replace('@', '_at_').replace('.', '_')
    return os.path.join(DATA_DIR, safe, "sips.json")


def load_sips():
    if USE_SUPABASE:
        res = _sb.table('sips').select('data').eq('user_email', _get_email()).execute()
        return [r['data'] for r in res.data]
    path = _local_sips_path()
    return json.load(open(path)) if os.path.exists(path) else []


def save_sips(sips):
    if USE_SUPABASE:
        email = _get_email()
        _sb.table('sips').delete().eq('user_email', email).execute()
        if sips:
            _sb.table('sips').insert(
                [{'id': s['id'], 'user_email': email, 'data': s} for s in sips]
            ).execute()
        return
    path = _local_sips_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(sips, f, indent=2)


# ---------------------------------------------------------------------------
# Owner / role-based access
# ---------------------------------------------------------------------------

def is_owner():
    if not OAUTH_CONFIGURED:
        return True
    if not OWNER_EMAIL:
        return True
    return (session.get('user') or {}).get('email', '') == OWNER_EMAIL


def owner_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not OAUTH_CONFIGURED:
            return f(*args, **kwargs)
        if not is_owner():
            return jsonify({"error": "Owner access required"}), 403
        return f(*args, **kwargs)
    return decorated


# ---------------------------------------------------------------------------
# Stocks data helpers
# ---------------------------------------------------------------------------

def _local_stocks_path():
    safe = _get_email().replace('@', '_at_').replace('.', '_')
    return os.path.join(DATA_DIR, safe, "stocks.json")


def load_stocks():
    if USE_SUPABASE:
        res = _sb.table('stocks').select('data').eq('user_email', _get_email()).execute()
        return [r['data'] for r in res.data]
    path = _local_stocks_path()
    return json.load(open(path)) if os.path.exists(path) else []


def save_stocks(stocks):
    if USE_SUPABASE:
        email = _get_email()
        _sb.table('stocks').delete().eq('user_email', email).execute()
        if stocks:
            _sb.table('stocks').insert(
                [{'id': s['id'], 'user_email': email, 'data': s} for s in stocks]
            ).execute()
        return
    path = _local_stocks_path()
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, "w") as f:
        json.dump(stocks, f, indent=2)


# ---------------------------------------------------------------------------
# Live stock prices via yfinance (NSE)
# ---------------------------------------------------------------------------

def _fetch_one_price(sym):
    sym_upper = sym.strip().upper()

    # Check cache
    if USE_SUPABASE:
        res = _sb.table('stock_cache').select('data, cached_at').eq('symbol', sym_upper).execute()
        if res.data and (time.time() - res.data[0]['cached_at']) < STOCK_CACHE_TTL:
            return sym_upper, res.data[0]['data']
    else:
        cache_path = os.path.join(CACHE_DIR, f"stk_{sym_upper}.json")
        if os.path.exists(cache_path) and (time.time() - os.path.getmtime(cache_path)) < STOCK_CACHE_TTL:
            return sym_upper, json.load(open(cache_path))

    try:
        import yfinance as yf
        t    = yf.Ticker(sym_upper + ".NS")
        hist = t.history(period="1y")
        if hist.empty:
            return sym_upper, None
        current   = float(hist["Close"].iloc[-1])
        high52    = float(hist["Close"].max())
        low52     = float(hist["Close"].min())
        range_pct = ((current - low52) / (high52 - low52) * 100) if (high52 - low52) > 0 else 50
        result    = {
            "symbol":        sym_upper,
            "current_price": round(current, 2),
            "week52_high":   round(high52, 2),
            "week52_low":    round(low52, 2),
            "range_pct":     round(range_pct, 1),
        }
        if USE_SUPABASE:
            _sb.table('stock_cache').upsert({'symbol': sym_upper, 'data': result, 'cached_at': time.time()}).execute()
        else:
            os.makedirs(CACHE_DIR, exist_ok=True)
            with open(os.path.join(CACHE_DIR, f"stk_{sym_upper}.json"), "w") as f:
                json.dump(result, f)
        return sym_upper, result
    except Exception:
        return sym_upper, None


def fetch_all_stock_prices(symbols):
    if not symbols:
        return {}
    results = {}
    with ThreadPoolExecutor(max_workers=6) as ex:
        futures = {ex.submit(_fetch_one_price, s): s for s in symbols}
        for fut in as_completed(futures):
            sym, data = fut.result()
            if data:
                results[sym] = data
    return results


# ---------------------------------------------------------------------------
# Stock recommendation engine
# ---------------------------------------------------------------------------

def stock_recommend(stock, price_data):
    cost = float(stock.get("avg_cost") or 0)
    qty  = float(stock.get("qty")      or 0)

    if not price_data:
        return {
            "status": "NO DATA", "color": "secondary", "icon": "❓",
            "action": "Could not fetch live price — verify the NSE symbol.",
            "pnl_pct": 0, "pnl_abs": 0, "invested": 0, "current_value": 0, "range_pct": 0,
        }

    current   = price_data["current_price"]
    range_pct = price_data.get("range_pct", 50)
    invested  = cost * qty
    curr_val  = current * qty
    pnl_pct   = ((current - cost) / cost * 100) if cost > 0 else 0
    pnl_abs   = curr_val - invested

    if cost <= 0:
        status, color, icon = "HOLD", "info", "👍"
        action = "Enter your average cost to get a personalised recommendation."
    elif pnl_pct < -25:
        status, color, icon = "EXIT", "dark", "🚨"
        action = f"Down {abs(pnl_pct):.1f}% — significant loss. Review fundamentals urgently."
    elif pnl_pct < -10:
        status, color, icon = "REVIEW", "danger", "⚠️"
        action = f"Down {abs(pnl_pct):.1f}% from cost. Assess if the thesis still holds."
    elif range_pct > 80:
        status, color, icon = "BOOK PARTIAL", "warning", "📤"
        action = f"Near 52-week high ({range_pct:.0f}% of range). Consider booking 25–30% profits."
    elif range_pct < 25 and pnl_pct > -8:
        status, color, icon = "ACCUMULATE", "success", "✅"
        action = f"Near 52-week low ({range_pct:.0f}% of range). Good opportunity to accumulate."
    elif pnl_pct > 50 and range_pct > 65:
        status, color, icon = "HOLD/TRIM", "info", "✂️"
        action = f"Up {pnl_pct:.1f}% — strong gains. Hold; consider trimming on a further rally."
    else:
        status, color, icon = "HOLD", "info", "👍"
        action = f"{'Up' if pnl_pct >= 0 else 'Down'} {abs(pnl_pct):.1f}% from cost. Continue holding."

    return {
        "status": status, "color": color, "icon": icon, "action": action,
        "pnl_pct":       round(pnl_pct, 2),
        "pnl_abs":       round(pnl_abs, 2),
        "invested":      round(invested, 2),
        "current_value": round(curr_val, 2),
        "range_pct":     round(range_pct, 1),
    }


# ---------------------------------------------------------------------------
# Fund NAV cache — Supabase in production, files locally
# ---------------------------------------------------------------------------

def fetch_fund_data(scheme_code):
    if not scheme_code:
        return None

    # Check cache
    if USE_SUPABASE:
        res = _sb.table('nav_cache').select('data, cached_at').eq('scheme_code', scheme_code).execute()
        if res.data and (time.time() - res.data[0]['cached_at']) < CACHE_TTL:
            return res.data[0]['data']
    else:
        path = os.path.join(CACHE_DIR, f"{scheme_code}.json")
        if os.path.exists(path) and (time.time() - os.path.getmtime(path)) < CACHE_TTL:
            return json.load(open(path))

    try:
        resp = requests.get(f"https://api.mfapi.in/mf/{scheme_code}", timeout=15)
        if resp.status_code == 200:
            data = resp.json()
            if USE_SUPABASE:
                _sb.table('nav_cache').upsert({'scheme_code': scheme_code, 'data': data, 'cached_at': time.time()}).execute()
            else:
                os.makedirs(CACHE_DIR, exist_ok=True)
                with open(os.path.join(CACHE_DIR, f"{scheme_code}.json"), "w") as f:
                    json.dump(data, f)
            return data
    except Exception:
        pass
    return None


def search_funds(query):
    try:
        resp = requests.get("https://api.mfapi.in/mf/search", params={"q": query}, timeout=10)
        if resp.status_code == 200:
            return resp.json()
    except Exception:
        pass
    return []


# ---------------------------------------------------------------------------
# NAV lookup (manual SIPs)
# ---------------------------------------------------------------------------

def build_nav_dict(nav_data):
    return {item["date"]: float(item["nav"]) for item in nav_data}


def get_nav_on_date(nav_dict, target_date):
    for offset in range(10):
        d = target_date - relativedelta(days=offset)
        key = d.strftime("%d-%m-%Y")
        if key in nav_dict:
            return nav_dict[key]
    return None


# ---------------------------------------------------------------------------
# XIRR calculation (manual SIPs)
# ---------------------------------------------------------------------------

def calc_xirr(cashflows):
    if len(cashflows) < 2:
        return 0.0
    dates, amounts = zip(*cashflows)
    t0 = min(dates)
    days = [(d - t0).days for d in dates]

    def npv(r):
        return sum(a / (1 + r) ** (d / 365.0) for a, d in zip(amounts, days))

    try:
        return brentq(npv, -0.9999, 100.0, maxiter=1000)
    except Exception:
        return 0.0


# ---------------------------------------------------------------------------
# Performance calculation
# ---------------------------------------------------------------------------

def calc_performance(sip, fund_data):
    # ── Groww-imported fund: use their data directly ──
    if sip.get("import_source") == "groww_excel":
        invested = float(sip.get("groww_invested") or 0)
        current  = float(sip.get("groww_current") or 0)
        units    = float(sip.get("groww_units") or 0)
        xirr_val = float(sip.get("groww_xirr") or 0)
        pl       = current - invested
        abs_ret  = (pl / invested * 100) if invested else 0
        current_nav = None
        if fund_data and fund_data.get("data"):
            current_nav = float(fund_data["data"][0]["nav"])
        return {
            "total_invested":   round(invested, 2),
            "current_value":    round(current, 2),
            "total_units":      round(units, 4),
            "current_nav":      round(current_nav, 4) if current_nav else None,
            "xirr":             round(xirr_val, 2),
            "absolute_return":  round(abs_ret, 2),
            "profit_loss":      round(pl, 2),
            "installments":     None,
            "duration_months":  24,   # safe default so recommendation runs
            "source":           "groww",
        }

    # ── Manual SIP: calculate from NAV history ──
    if not fund_data:
        return None
    nav_history = fund_data.get("data", [])
    if not nav_history:
        return None

    nav_dict    = build_nav_dict(nav_history)
    current_nav = float(nav_history[0]["nav"])

    start_date = datetime.strptime(sip["start_date"], "%Y-%m-%d").date()
    today      = date.today()
    amount     = float(sip["amount"])

    cashflows       = []
    total_units     = 0.0
    total_invested  = 0.0
    installment_count = 0

    cur = start_date
    while cur <= today:
        nav = get_nav_on_date(nav_dict, cur)
        if nav and nav > 0:
            units = amount / nav
            total_units    += units
            total_invested += amount
            cashflows.append((cur, -amount))
            installment_count += 1
        cur += relativedelta(months=1)

    if total_invested == 0:
        return None

    current_value = total_units * current_nav
    cashflows.append((today, current_value))

    xirr_val   = calc_xirr(cashflows) * 100
    abs_return = (current_value - total_invested) / total_invested * 100
    profit_loss = current_value - total_invested
    months = (today.year - start_date.year) * 12 + (today.month - start_date.month)

    return {
        "total_invested":  round(total_invested, 2),
        "current_value":   round(current_value, 2),
        "total_units":     round(total_units, 4),
        "current_nav":     round(current_nav, 4),
        "xirr":            round(xirr_val, 2),
        "absolute_return": round(abs_return, 2),
        "profit_loss":     round(profit_loss, 2),
        "installments":    installment_count,
        "duration_months": months,
        "source":          "manual",
    }


# ---------------------------------------------------------------------------
# Recommendation engine
# ---------------------------------------------------------------------------

def recommend(sip, perf):
    xirr_val       = perf.get("xirr", 0)
    months         = perf.get("duration_months", 0)
    is_groww_import = perf.get("source") == "groww"

    raw_cat = sip.get("category", "equity").lower().strip()
    bench   = None
    for key, val in CATEGORY_BENCHMARKS.items():
        if key in raw_cat or raw_cat in key:
            bench = val
            break
    if bench is None:
        bench = CATEGORY_BENCHMARKS["equity"]

    target     = bench["xirr"]
    index_name = bench["index"]

    # For Groww imports with real XIRR we skip the TOO EARLY check
    if months < 12 and not is_groww_import:
        return {
            "status": "TOO EARLY", "color": "secondary", "icon": "🕐",
            "reason": f"Only {months} month(s) of data — need at least 12 months to judge.",
            "action": "Continue SIP, revisit after a year.",
            "benchmark": target, "index": index_name,
            "gap": round(xirr_val - target, 2),
        }

    gap = xirr_val - target

    if gap >= 3:
        status, color, icon = "KEEP", "success", "✅"
        reason = f"Outperforming {index_name} benchmark ({target}%) by {gap:.1f}%."
        action = "Strong performer — continue SIP confidently."
    elif gap >= -2:
        status, color, icon = "KEEP", "info", "👍"
        reason = f"Performing in line with {index_name} benchmark ({target}%). Gap: {gap:+.1f}%."
        action = "Acceptable — monitor quarterly and continue."
    elif gap >= -5:
        status, color, icon = "REVIEW", "warning", "⚠️"
        reason = f"Underperforming {index_name} benchmark ({target}%) by {abs(gap):.1f}%."
        action = "Review fund fundamentals. Consider switching if trend continues 2+ quarters."
    elif xirr_val >= 0:
        status, color, icon = "DILUTE", "danger", "🔻"
        reason = f"Significantly underperforming {index_name} (target {target}%) by {abs(gap):.1f}%."
        action = "Pause new SIPs. Reinvest in a better fund in the same category."
    else:
        status, color, icon = "EXIT", "dark", "🚨"
        reason = f"Negative XIRR ({xirr_val:.1f}%) — capital destruction despite long investment."
        action = "Stop SIP immediately. Evaluate exit and redeploy capital."

    return {
        "status": status, "color": color, "icon": icon,
        "reason": reason, "action": action,
        "benchmark": target, "index": index_name,
        "gap": round(gap, 2),
    }


# ---------------------------------------------------------------------------
# Groww MF Excel parser
# ---------------------------------------------------------------------------

def parse_groww_excel(file_bytes):
    """Parse Groww Holdings .xlsx and return list of fund dicts."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open Excel file: {e}")

    ws = wb.active

    # Find the header row (contains "Scheme Name")
    header_row_idx = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and row[0] == "Scheme Name":
            header_row_idx = i
            break

    if not header_row_idx:
        raise ValueError("Could not find 'Scheme Name' header row in the Excel file.")

    # Build column index map from header row
    headers = [ws.cell(header_row_idx, j).value for j in range(1, ws.max_column + 1)]
    col = {}
    for j, h in enumerate(headers):
        if h and h not in col:
            col[h] = j  # 0-based index into row list

    required = ["Scheme Name", "Invested Value", "Current Value", "XIRR"]
    for r in required:
        if r not in col:
            raise ValueError(f"Missing expected column '{r}' in Excel file.")

    def cell(row, name, default=None):
        idx = col.get(name)
        return row[idx] if idx is not None and idx < len(row) else default

    def parse_float(v, default=0.0):
        if v is None:
            return default
        if isinstance(v, (int, float)):
            return float(v)
        return float(str(v).replace(",", "").replace("%", "").strip() or default)

    def parse_xirr(v):
        """Convert XIRR cell to float percentage (e.g. '11.01%' → 11.01)."""
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            # Sometimes stored as decimal (0.1101) instead of percentage
            f = float(v)
            return round(f * 100 if abs(f) < 2 else f, 2)
        s = str(v).replace("%", "").strip()
        return float(s) if s else 0.0

    funds = []
    # Data starts 2 rows after header (blank separator row in Groww format)
    data_start = header_row_idx + 2

    for i in range(data_start, ws.max_row + 1):
        row = [ws.cell(i, j).value for j in range(1, ws.max_column + 1)]
        scheme_name = str(cell(row, "Scheme Name") or "").strip()
        if not scheme_name:
            continue

        sub_cat_raw = str(cell(row, "Sub-category") or "").strip()
        category = SUBCATEGORY_MAP.get(sub_cat_raw.lower(), "equity")

        invested = parse_float(cell(row, "Invested Value"))
        current  = parse_float(cell(row, "Current Value"))

        fund = {
            "fund_name":    scheme_name,
            "amc":          str(cell(row, "AMC") or "").strip(),
            "groww_category": str(cell(row, "Category") or "").strip(),
            "sub_category": sub_cat_raw,
            "category":     category,
            "folio_no":     str(cell(row, "Folio No.") or "").strip(),
            "groww_source": str(cell(row, "Source") or "").strip(),
            "groww_units":  parse_float(cell(row, "Units")),
            "groww_invested": invested,
            "groww_current":  current,
            "groww_returns":  parse_float(cell(row, "Returns")),
            "groww_xirr":     parse_xirr(cell(row, "XIRR")),
            "import_source":  "groww_excel",
            "groww_last_import": date.today().isoformat(),
        }
        funds.append(fund)

    return funds


# ---------------------------------------------------------------------------
# Groww Stocks/Equity Excel parser
# ---------------------------------------------------------------------------

def _resolve_isin_to_symbol(isin):
    """Resolve an Indian ISIN to its NSE ticker via Yahoo Finance search."""
    try:
        url = (
            f'https://query2.finance.yahoo.com/v1/finance/search'
            f'?q={isin}&lang=en-US&region=IN&quotesCount=6&newsCount=0'
        )
        r = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'}, timeout=6)
        if r.status_code == 200:
            for q in r.json().get('quotes', []):
                sym = q.get('symbol', '')
                if sym.endswith('.NS'):
                    return sym[:-3]          # strip .NS suffix
    except Exception:
        pass
    return None


def parse_groww_stocks_excel(file_bytes):
    """Parse Groww Holdings Statement .xlsx and return list of stock dicts.

    Handles the format:  Stock Name | ISIN | Quantity | Average buy price |
                         Buy value  | Closing price | Closing value | Unrealised P&L
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Cannot open Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Find header row: must contain both 'isin' and 'quantity' (case-insensitive)
    header_idx = None
    for i, row in enumerate(rows):
        cells = [str(c).strip().lower() for c in row if c is not None]
        if 'isin' in cells and any('quantity' in c or 'qty' in c for c in cells):
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(
            "Could not find a header row with ISIN and Quantity columns. "
            "Please upload a Groww Holdings Statement export."
        )

    headers = [str(v).strip().lower() if v is not None else '' for v in rows[header_idx]]

    def col_of(*candidates):
        for cand in candidates:
            for j, h in enumerate(headers):
                if h and cand in h:
                    return j
        return None

    name_col     = col_of('stock name', 'company', 'name', 'scrip')
    isin_col     = col_of('isin')
    qty_col      = col_of('quantity', 'qty')
    avg_col      = col_of('average buy price', 'avg price', 'average price', 'buy price')
    invested_col = col_of('buy value', 'invested', 'cost value')
    ltp_col      = col_of('closing price', 'ltp', 'cmp', 'current price')
    curr_val_col = col_of('closing value', 'current value', 'market value')
    pnl_col      = col_of('unrealised', 'p&l', 'profit', 'gain')

    if isin_col is None or qty_col is None:
        raise ValueError(
            "Could not find required columns (ISIN, Quantity). "
            "Please check this is a Groww equity holdings export."
        )

    def to_float(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(',', '').replace('₹', '').replace('%', '').strip()
        try:
            return float(s)
        except ValueError:
            return None

    today = date.today().isoformat()
    stocks = []

    for row in rows[header_idx + 1:]:
        if not any(v is not None for v in row):
            continue

        isin = str(row[isin_col] or '').strip().upper()
        if not isin or not isin.startswith('IN') or len(isin) != 12:
            continue

        qty = to_float(row[qty_col] if qty_col is not None else None)
        if not qty or qty <= 0:
            continue

        name      = str(row[name_col] or '').strip() if name_col is not None else isin
        avg_cost  = to_float(row[avg_col]      if avg_col is not None else None)
        invested  = to_float(row[invested_col] if invested_col is not None else None)
        ltp       = to_float(row[ltp_col]      if ltp_col is not None else None)
        curr_val  = to_float(row[curr_val_col] if curr_val_col is not None else None)
        pnl       = to_float(row[pnl_col]      if pnl_col is not None else None)

        # avg_cost = 0 means Groww doesn't know the buy price (demat transfer etc.)
        if avg_cost == 0.0:
            avg_cost = None
        if invested == 0.0:
            invested = None

        pnl_pct = None
        if pnl is not None and invested:
            pnl_pct = round(pnl / invested * 100, 2)

        stocks.append({
            "isin":                 isin,
            "symbol":               None,     # resolved later via Yahoo Finance
            "company_name":         name,
            "qty":                  qty,
            "avg_cost":             avg_cost,
            "groww_ltp":            ltp,
            "groww_current_value":  curr_val,
            "groww_invested":       invested,
            "groww_pnl":            pnl,
            "groww_pnl_pct":        pnl_pct,
            "import_source":        "groww_stocks_excel",
            "groww_last_import":    today,
        })

    return stocks


# ---------------------------------------------------------------------------
# Flask routes
# ---------------------------------------------------------------------------

@app.route("/")
@login_required
def index():
    return render_template("index.html", user=session.get('user'), is_owner=is_owner())


@app.route("/api/sips", methods=["GET"])
@login_required
def api_get_sips():
    return jsonify(load_sips())


@app.route("/api/sips", methods=["POST"])
@login_required
def api_add_sip():
    sips = load_sips()
    sip  = request.json
    existing_ids = {int(s.get("id", 0)) for s in sips if str(s.get("id", "")).isdigit()}
    sip["id"]         = str(max(existing_ids, default=0) + 1)
    sip["created_at"] = datetime.now().isoformat()
    sips.append(sip)
    save_sips(sips)
    return jsonify(sip)


@app.route("/api/sips/<sip_id>", methods=["PUT"])
@login_required
def api_update_sip(sip_id):
    sips = load_sips()
    body = request.json
    for i, s in enumerate(sips):
        if s["id"] == sip_id:
            sips[i] = {**s, **body, "id": sip_id}
            save_sips(sips)
            return jsonify(sips[i])
    return jsonify({"error": "Not found"}), 404


@app.route("/api/sips/<sip_id>", methods=["DELETE"])
@login_required
def api_delete_sip(sip_id):
    sips = [s for s in load_sips() if s["id"] != sip_id]
    save_sips(sips)
    return jsonify({"success": True})


@app.route("/api/import/groww", methods=["POST"])
@login_required
def api_import_groww():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an Excel (.xlsx) file"}), 400

    try:
        parsed = parse_groww_excel(f.read())
    except ValueError as e:
        return jsonify({"error": str(e)}), 400

    if not parsed:
        return jsonify({"error": "No fund data found in the file"}), 400

    sips = load_sips()

    # Composite key: folio_no + fund_name (same folio can hold multiple schemes)
    def _import_key(s):
        return f"{s.get('folio_no', '')}|{s.get('fund_name', '')}"

    folio_index = {_import_key(s): i for i, s in enumerate(sips) if s.get("folio_no")}
    existing_ids = {int(s.get("id", 0)) for s in sips if str(s.get("id", "")).isdigit()}
    next_id = max(existing_ids, default=0) + 1

    added = updated = 0

    for fund in parsed:
        key = _import_key(fund)

        if key in folio_index:
            # Update Groww fields; preserve user's label/notes/scheme_code
            idx = folio_index[key]
            sips[idx] = {
                **sips[idx],
                "groww_units":       fund["groww_units"],
                "groww_invested":    fund["groww_invested"],
                "groww_current":     fund["groww_current"],
                "groww_returns":     fund["groww_returns"],
                "groww_xirr":        fund["groww_xirr"],
                "groww_last_import": fund["groww_last_import"],
                "import_source":     "groww_excel",
                # Update category only if not already customised
                "category": sips[idx].get("category") or fund["category"],
                "sub_category": fund["sub_category"],
                "groww_category": fund["groww_category"],
                "amc": fund["amc"],
            }
            updated += 1
        else:
            # New fund — create clean label from scheme name
            label = (fund["fund_name"]
                     .replace(" Direct Growth", "")
                     .replace(" Direct Plan Growth", "")
                     .replace(" Direct Plan", "")
                     .strip())
            label = label[:50]

            fund["id"]         = str(next_id)
            fund["label"]      = label
            fund["amount"]     = 0       # unknown for imports
            fund["start_date"] = None
            fund["created_at"] = datetime.now().isoformat()
            next_id += 1

            sips.append(fund)
            folio_index[key] = len(sips) - 1
            added += 1

    save_sips(sips)
    return jsonify({
        "success": True,
        "added":   added,
        "updated": updated,
        "total":   len(parsed),
    })


@app.route("/api/dashboard")
@login_required
def api_dashboard():
    sips    = load_sips()
    results = []

    for sip in sips:
        is_import = sip.get("import_source") == "groww_excel"

        # Manual SIPs need a scheme_code; imports don't
        if not sip.get("scheme_code") and not is_import:
            results.append({**sip, "error": "No scheme code", "performance": None, "recommendation": None})
            continue

        fund_data = fetch_fund_data(sip.get("scheme_code")) if sip.get("scheme_code") else None

        if not fund_data and not is_import:
            results.append({**sip, "error": "Could not fetch fund data", "performance": None, "recommendation": None})
            continue

        meta = fund_data.get("meta", {}) if fund_data else {}
        perf = calc_performance(sip, fund_data)
        rec  = recommend(sip, perf) if perf else None

        results.append({
            **sip,
            "fund_name_api":   meta.get("fund_name", ""),
            "scheme_type":     meta.get("scheme_type", ""),
            "scheme_category": meta.get("scheme_category", ""),
            "fund_house":      meta.get("fund_house", "") or sip.get("amc", ""),
            "performance":     perf,
            "recommendation":  rec,
            "error":           None,
        })

    valid         = [r for r in results if r["performance"]]
    total_invested = sum(r["performance"]["total_invested"] for r in valid)
    total_current  = sum(r["performance"]["current_value"]  for r in valid)
    total_pl       = total_current - total_invested
    overall_return = (total_pl / total_invested * 100) if total_invested else 0

    status_counts = {}
    for r in valid:
        s = r["recommendation"]["status"] if r.get("recommendation") else "N/A"
        status_counts[s] = status_counts.get(s, 0) + 1

    return jsonify({
        "sips": results,
        "summary": {
            "total_invested":     round(total_invested, 2),
            "total_current":      round(total_current, 2),
            "total_pl":           round(total_pl, 2),
            "overall_return_pct": round(overall_return, 2),
            "count":   len(results),
            "valid":   len(valid),
            "status_counts": status_counts,
            "last_updated": datetime.now().strftime("%d %b %Y, %I:%M %p"),
        },
    })


@app.route("/api/search-funds")
@login_required
def api_search_funds():
    q = request.args.get("q", "").strip()
    if len(q) < 3:
        return jsonify([])
    return jsonify(search_funds(q)[:20])


@app.route("/api/clear-cache", methods=["POST"])
@login_required
def api_clear_cache():
    if USE_SUPABASE:
        _sb.table('nav_cache').delete().neq('scheme_code', '___').execute()
        return jsonify({"cleared": "all"})
    count = 0
    if os.path.exists(CACHE_DIR):
        for f in os.listdir(CACHE_DIR):
            if f.endswith(".json"):
                os.remove(os.path.join(CACHE_DIR, f))
                count += 1
    return jsonify({"cleared": count})


# ---------------------------------------------------------------------------
# Stocks routes (owner only)
# ---------------------------------------------------------------------------

@app.route("/api/stocks", methods=["GET"])
@login_required
def api_get_stocks():
    return jsonify(load_stocks())


@app.route("/api/stocks", methods=["POST"])
@login_required
def api_add_stock():
    stocks = load_stocks()
    stock  = request.json
    existing_ids = {int(s.get("id", 0)) for s in stocks if str(s.get("id", "")).isdigit()}
    stock["id"]         = str(max(existing_ids, default=0) + 1)
    stock["created_at"] = datetime.now().isoformat()
    stocks.append(stock)
    save_stocks(stocks)
    return jsonify(stock)


@app.route("/api/stocks/<stock_id>", methods=["PUT"])
@login_required
def api_update_stock(stock_id):
    stocks = load_stocks()
    body   = request.json
    for i, s in enumerate(stocks):
        if s["id"] == stock_id:
            stocks[i] = {**s, **body, "id": stock_id}
            save_stocks(stocks)
            return jsonify(stocks[i])
    return jsonify({"error": "Not found"}), 404


@app.route("/api/stocks/<stock_id>", methods=["DELETE"])
@login_required
def api_delete_stock(stock_id):
    stocks = [s for s in load_stocks() if s["id"] != stock_id]
    save_stocks(stocks)
    return jsonify({"success": True})


@app.route("/api/import/groww-stocks", methods=["POST"])
@login_required
def api_import_groww_stocks():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith((".xlsx", ".xls")):
        return jsonify({"error": "Please upload an Excel (.xlsx) file"}), 400
    try:
        parsed = parse_groww_stocks_excel(f.read())
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    if not parsed:
        return jsonify({"error": "No stock data found in the file"}), 400

    # Resolve ISINs → NSE symbols in parallel (needed for yfinance live prices)
    isins_to_resolve = [s["isin"] for s in parsed if s.get("symbol") is None]
    isin_symbol_map  = {}
    if isins_to_resolve:
        with ThreadPoolExecutor(max_workers=8) as ex:
            futures = {ex.submit(_resolve_isin_to_symbol, isin): isin for isin in isins_to_resolve}
            for fut in as_completed(futures):
                isin = futures[fut]
                try:
                    sym = fut.result()
                    if sym:
                        isin_symbol_map[isin] = sym
                except Exception:
                    pass

    for stock in parsed:
        if stock.get("symbol") is None:
            stock["symbol"] = isin_symbol_map.get(stock["isin"]) or stock["isin"]

    stocks = load_stocks()
    # Build lookup indexes
    isin_index   = {s.get("isin", "").upper(): i for i, s in enumerate(stocks) if s.get("isin")}
    symbol_index = {s.get("symbol", "").upper(): i for i, s in enumerate(stocks) if s.get("symbol")}
    existing_ids = {int(s.get("id", 0)) for s in stocks if str(s.get("id", "")).isdigit()}
    next_id = max(existing_ids, default=0) + 1
    added = updated = 0

    for stock in parsed:
        isin = stock.get("isin", "").upper()
        sym  = (stock.get("symbol") or "").upper()

        # Match by ISIN first, then fall back to symbol
        idx = isin_index.get(isin) if isin else None
        if idx is None and sym:
            idx = symbol_index.get(sym)

        if idx is not None:
            stocks[idx] = {
                **stocks[idx],
                "isin":                 stock.get("isin"),
                "symbol":               stock.get("symbol") or stocks[idx].get("symbol"),
                "qty":                  stock["qty"],
                "avg_cost":             stock["avg_cost"],
                "company_name":         stock.get("company_name") or stocks[idx].get("company_name"),
                "groww_ltp":            stock.get("groww_ltp"),
                "groww_current_value":  stock.get("groww_current_value"),
                "groww_invested":       stock.get("groww_invested"),
                "groww_pnl":            stock.get("groww_pnl"),
                "groww_pnl_pct":        stock.get("groww_pnl_pct"),
                "groww_last_import":    stock["groww_last_import"],
                "import_source":        "groww_stocks_excel",
            }
            # Update indexes with any newly resolved info
            if stock.get("isin"):
                isin_index[stock["isin"].upper()] = idx
            if stock.get("symbol"):
                symbol_index[stock["symbol"].upper()] = idx
            updated += 1
        else:
            stock["id"]         = str(next_id)
            stock["created_at"] = datetime.now().isoformat()
            stocks.append(stock)
            if isin:
                isin_index[isin] = len(stocks) - 1
            if sym:
                symbol_index[sym] = len(stocks) - 1
            next_id += 1
            added += 1

    save_stocks(stocks)
    return jsonify({"success": True, "added": added, "updated": updated, "total": len(parsed)})


# ---------------------------------------------------------------------------
# Feature 5: Nifty 50 benchmark
# ---------------------------------------------------------------------------

def fetch_nifty_data():
    try:
        import yfinance as yf
        t = yf.Ticker("^NSEI")
        hist = t.history(period="366d")
        if len(hist) >= 50:
            current = float(hist['Close'].iloc[-1])
            year_ago = float(hist['Close'].iloc[0])
            return_1y = round((current - year_ago) / year_ago * 100, 2)
            return {"current": round(current, 2), "return_1y": return_1y}
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Feature 6: Tax P&L summary
# ---------------------------------------------------------------------------

def calc_tax_summary(stocks_with_recs):
    today = date.today()
    stcg_positions, ltcg_positions = [], []
    for r in stocks_with_recs:
        pnl = r["recommendation"].get("pnl_abs", 0)
        invested = r["recommendation"].get("invested", 0)
        if pnl <= 0 or invested <= 0:
            continue
        try:
            buy_date = datetime.fromisoformat(r.get("created_at", "")).date()
        except Exception:
            continue
        holding_days = (today - buy_date).days
        entry = {
            "name": r.get("company_name") or r.get("symbol", ""),
            "symbol": r.get("symbol", ""),
            "pnl": round(pnl, 2),
            "holding_days": holding_days,
            "pnl_pct": r["recommendation"].get("pnl_pct", 0),
        }
        if holding_days < 365:
            stcg_positions.append(entry)
        else:
            ltcg_positions.append(entry)
    stcg_total = sum(p["pnl"] for p in stcg_positions)
    ltcg_total = sum(p["pnl"] for p in ltcg_positions)
    stcg_tax = round(stcg_total * 0.20, 2)
    ltcg_exempt = 125000
    ltcg_taxable = max(0, ltcg_total - ltcg_exempt)
    ltcg_tax = round(ltcg_taxable * 0.125, 2)
    return {
        "stcg_profit": round(stcg_total, 2),
        "stcg_tax": stcg_tax,
        "ltcg_profit": round(ltcg_total, 2),
        "ltcg_tax": ltcg_tax,
        "ltcg_exempt": ltcg_exempt,
        "total_tax": round(stcg_tax + ltcg_tax, 2),
        "stcg_positions": sorted(stcg_positions, key=lambda x: -x["pnl"]),
        "ltcg_positions": sorted(ltcg_positions, key=lambda x: -x["pnl"]),
    }


# ---------------------------------------------------------------------------
# Feature 7: Portfolio history snapshots
# NOTE: Run the following SQL in Supabase to create the required table:
#
# CREATE TABLE portfolio_snapshots (
#   user_email TEXT NOT NULL,
#   snapshot_date DATE NOT NULL,
#   mf_invested FLOAT DEFAULT 0,
#   mf_current FLOAT DEFAULT 0,
#   stocks_invested FLOAT DEFAULT 0,
#   stocks_current FLOAT DEFAULT 0,
#   PRIMARY KEY (user_email, snapshot_date)
# );
# ---------------------------------------------------------------------------

def load_snapshots():
    if USE_SUPABASE:
        res = _sb.table('portfolio_snapshots').select('*').eq('user_email', _get_email()).order('snapshot_date').execute()
        return res.data
    path = os.path.join(DATA_DIR, _get_email().replace('@', '_at_').replace('.', '_'), 'snapshots.json')
    return json.load(open(path)) if os.path.exists(path) else []


def save_snapshot(snap):
    if USE_SUPABASE:
        _sb.table('portfolio_snapshots').upsert({
            'user_email': _get_email(),
            'snapshot_date': snap['snapshot_date'],
            'mf_invested': snap.get('mf_invested', 0),
            'mf_current': snap.get('mf_current', 0),
            'stocks_invested': snap.get('stocks_invested', 0),
            'stocks_current': snap.get('stocks_current', 0),
        }).execute()
        return
    safe = _get_email().replace('@', '_at_').replace('.', '_')
    path = os.path.join(DATA_DIR, safe, 'snapshots.json')
    os.makedirs(os.path.dirname(path), exist_ok=True)
    snaps = json.load(open(path)) if os.path.exists(path) else []
    # upsert by date
    snaps = [s for s in snaps if s.get('snapshot_date') != snap['snapshot_date']]
    snaps.append(snap)
    snaps.sort(key=lambda x: x['snapshot_date'])
    with open(path, 'w') as f:
        json.dump(snaps, f, indent=2)


@app.route("/api/portfolio/snapshot", methods=["POST"])
@login_required
def api_portfolio_snapshot():
    body = request.json or {}
    today = date.today().isoformat()
    save_snapshot({
        "snapshot_date":   today,
        "mf_invested":     body.get("mf_invested", 0),
        "mf_current":      body.get("mf_current", 0),
        "stocks_invested": body.get("stocks_invested", 0),
        "stocks_current":  body.get("stocks_current", 0),
    })
    return jsonify({"success": True, "date": today})


@app.route("/api/portfolio/history")
@login_required
def api_portfolio_history():
    snaps = load_snapshots()
    return jsonify({"snapshots": snaps[-90:]})  # last 90 days


@app.route("/api/stocks/dashboard")
@login_required
def api_stocks_dashboard():
    stocks  = load_stocks()
    symbols = list({s.get("symbol", "").upper() for s in stocks if s.get("symbol")})
    prices  = fetch_all_stock_prices(symbols)

    results = []
    for stock in stocks:
        sym = stock.get("symbol", "").upper()
        pd_ = prices.get(sym)
        rec = stock_recommend(stock, pd_)
        results.append({**stock, "price_data": pd_, "recommendation": rec})

    total_invested = 0.0
    total_current  = 0.0
    total_pnl      = 0.0

    for r in results:
        rec = r["recommendation"]
        inv = rec.get("invested") or 0

        # Current value: prefer live NSE price, fall back to Groww closing value
        live_curr = rec.get("current_value") or 0
        curr = live_curr if live_curr > 0 else (r.get("groww_current_value") or 0)

        total_current  += curr
        total_invested += inv

        # P&L: use calculated P&L where cost basis known, else Groww's own P&L
        if inv > 0:
            total_pnl += rec.get("pnl_abs") or 0
        else:
            total_pnl += r.get("groww_pnl") or 0

    overall_return = (total_pnl / total_invested * 100) if total_invested else 0

    status_counts = {}
    for r in results:
        s = r["recommendation"]["status"]
        status_counts[s] = status_counts.get(s, 0) + 1

    return jsonify({
        "stocks": results,
        "summary": {
            "total_invested":     round(total_invested, 2),
            "total_current":      round(total_current, 2),
            "total_pnl":          round(total_pnl, 2),
            "overall_return_pct": round(overall_return, 2),
            "count":              len(results),
            "status_counts":      status_counts,
            "last_updated":       datetime.now().strftime("%d %b %Y, %I:%M %p"),
            "nifty":              fetch_nifty_data(),
            "tax":                calc_tax_summary(results),
        }
    })


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5001))
    print(f"\n🚀  SIP Tracker → http://localhost:{port}\n")
    app.run(debug=False, port=port)
